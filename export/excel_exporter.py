# export/excel_exporter.py
"""Module xuất dữ liệu ra Excel"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import List, Dict
import io
from src.utils import format_number


class ExcelExporter:
    """Class xuất dữ liệu ra Excel"""
    
    def __init__(self):
        """Khởi tạo exporter"""
        self.wb = Workbook()
        self.ws = self.wb.active
        
        # Định nghĩa styles
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.title_font = Font(bold=True, size=14)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def create_payment_schedule_excel(self, schedule: List[Dict], 
                                     loan_info: Dict) -> io.BytesIO:
        """
        Tạo file Excel bảng kê kế hoạch trả nợ
        
        Args:
            schedule: Lịch trả nợ
            loan_info: Thông tin khoản vay
            
        Returns:
            BytesIO object chứa file Excel
        """
        # Tạo DataFrame
        df_schedule = pd.DataFrame(schedule)
        
        # Đổi tên cột
        df_schedule.columns = ['Kỳ', 'Gốc', 'Lãi', 'Tổng trả', 'Dư nợ']
        
        # Clear sheet
        self.ws.delete_rows(1, self.ws.max_row)
        
        # Tiêu đề
        self.ws['A1'] = 'BẢNG KÊ KẾ HOẠCH TRẢ NỢ VAY'
        self.ws['A1'].font = self.title_font
        self.ws.merge_cells('A1:E1')
        self.ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Thông tin khoản vay
        row = 3
        info_items = [
            ('Khách hàng:', loan_info.get('customer_name', 'N/A')),
            ('Số tiền vay:', f"{format_number(loan_info.get('loan_amount', 0))} VND"),
            ('Lãi suất:', f"{loan_info.get('interest_rate', 0)}% /năm"),
            ('Thời hạn:', f"{loan_info.get('loan_term', 0)} tháng"),
        ]
        
        for label, value in info_items:
            self.ws[f'A{row}'] = label
            self.ws[f'A{row}'].font = Font(bold=True)
            self.ws[f'B{row}'] = value
            row += 1
        
        # Bảng lịch trả nợ
        row += 1
        header_row = row
        
        # Header
        headers = ['Kỳ', 'Gốc (VND)', 'Lãi (VND)', 'Tổng trả (VND)', 'Dư nợ (VND)']
        for col, header in enumerate(headers, start=1):
            cell = self.ws.cell(row=header_row, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # Data
        for idx, period in enumerate(schedule, start=1):
            row = header_row + idx
            
            self.ws.cell(row=row, column=1, value=period['month'])
            self.ws.cell(row=row, column=2, value=period['principal'])
            self.ws.cell(row=row, column=3, value=period['interest'])
            self.ws.cell(row=row, column=4, value=period['total_payment'])
            self.ws.cell(row=row, column=5, value=period['remaining_balance'])
            
            # Format và border
            for col in range(1, 6):
                cell = self.ws.cell(row=row, column=col)
                cell.border = self.border
                cell.alignment = Alignment(horizontal='right' if col > 1 else 'center')
                
                # Format số
                if col > 1:
                    cell.number_format = '#,##0'
        
        # Tổng cộng
        row += 1
        self.ws.cell(row=row, column=1, value='TỔNG CỘNG').font = Font(bold=True)
        self.ws.cell(row=row, column=2, value=sum(p['principal'] for p in schedule))
        self.ws.cell(row=row, column=3, value=sum(p['interest'] for p in schedule))
        self.ws.cell(row=row, column=4, value=sum(p['total_payment'] for p in schedule))
        
        for col in range(1, 6):
            cell = self.ws.cell(row=row, column=col)
            cell.font = Font(bold=True)
            cell.border = self.border
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            if col > 1:
                cell.number_format = '#,##0'
        
        # Điều chỉnh độ rộng cột
        self.ws.column_dimensions['A'].width = 10
        for col in ['B', 'C', 'D', 'E']:
            self.ws.column_dimensions[col].width = 20
        
        # Lưu vào BytesIO
        output = io.BytesIO()
        self.wb.save(output)
        output.seek(0)
        return output
    
    def create_financial_summary_excel(self, data: Dict) -> io.BytesIO:
        """
        Tạo file Excel tóm tắt tài chính
        
        Args:
            data: Dữ liệu tài chính
            
        Returns:
            BytesIO object chứa file Excel
        """
        # Tạo workbook mới
        wb = Workbook()
        ws = wb.active
        ws.title = "Tóm tắt tài chính"
        
        # Tiêu đề
        ws['A1'] = 'BÁO CÁO TÓM TẮT THẨM ĐỊNH TÀI CHÍNH'
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:B1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Thông tin
        row = 3
        sections = {
            'THÔNG TIN KHÁCH HÀNG': [
                ('Họ tên:', data.get('customer_name', 'N/A')),
                ('CCCD:', data.get('customer_cccd', 'N/A')),
                ('Địa chỉ:', data.get('customer_address', 'N/A')),
                ('Điện thoại:', data.get('customer_phone', 'N/A')),
            ],
            'THÔNG TIN KHOẢN VAY': [
                ('Mục đích:', data.get('loan_purpose', 'N/A')),
                ('Số tiền vay:', f"{format_number(data.get('loan_amount', 0))} VND"),
                ('Lãi suất:', f"{data.get('interest_rate', 0)}% /năm"),
                ('Thời hạn:', f"{data.get('loan_term', 0)} tháng"),
                ('Trả nợ hàng tháng:', f"{format_number(data.get('monthly_payment', 0))} VND"),
            ],
            'CHỈ TIÊU TÀI CHÍNH': [
                ('Thu nhập tháng:', f"{format_number(data.get('monthly_income', 0))} VND"),
                ('Chi phí tháng:', f"{format_number(data.get('monthly_expense', 0))} VND"),
                ('Dòng tiền ròng:', f"{format_number(data.get('net_cash_flow', 0))} VND"),
                ('DSR:', f"{data.get('dsr', 0):.2f}%"),
                ('Biên an toàn:', f"{data.get('safety_margin', 0):.2f}%"),
            ],
            'ĐÁNH GIÁ': [
                ('Kết luận:', data.get('assessment', 'N/A')),
                ('Mức độ rủi ro:', data.get('risk_level', 'N/A')),
            ]
        }
        
        for section_title, items in sections.items():
            ws[f'A{row}'] = section_title
            ws[f'A{row}'].font = Font(bold=True, size=12)
            ws.merge_cells(f'A{row}:B{row}')
            row += 1
            
            for label, value in items:
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'] = value
                row += 1
            
            row += 1
        
        # Điều chỉnh độ rộng cột
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 50
        
        # Lưu vào BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
